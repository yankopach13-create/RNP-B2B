# src/features/dashboard.py
from __future__ import annotations
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Callable
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from config.constants import (
    MINSK_REGION_COMPONENTS,
    SPECIAL_RETAIL_SUBDIVISIONS,
)
from features.category_order import (
    COL_GENERAL_SPEC,
    COL_GENERAL_TRADITION,
    COL_SPEC_RNP,
    COL_TRADITION_RNP,
    COL_TURNOVER,
    _label_key,
    extract_category_row_values,
    get_category_source_column,
    get_razrez_source_column,
    load_category_order_list,
)
from features.data_prep import prepare_dataset
from features.factor_analysis import (
    RTRADE_CLIENTS,
    _build_factor_table,
    _build_category_rows,
    _build_segment_rows,
    _merge_categories,
    _prepare_factor_base,
)
from features.orders import _count_clients, _prepare_orders_dataset
from features.orders_metrics import calculate_orders_category_metrics
from features.render import (
    _format_money,
    _format_money_compact,
    _format_percent,
    _format_quantity,
    _format_quantity_compact,
    _table_height_from_rows,
    build_category_table,
    build_category_vertical_table,
    build_combined_finance_categories_table,
    build_financial_metrics_table,
    build_financial_metrics_vertical_table,
    build_turnover_table,
    calculate_turnover_by_category,
    render_section,
    render_turnover_block,
)
from features.factor_analysis import render_factor_analysis
from data.references import (
    REF_CATEGORIES,
    REF_CATEGORY_ORDER,
    REF_CONTRACTORS,
    REF_DZ_REMOVE,
    REF_DZ_SPEC,
    REF_DZ_TRAD,
    REF_SALES_POD_CARTRIDGE,
    append_reference_rows,
    get_reference_label,
    get_reference_title,
    load_reference,
)
from features.hardware_sales_dynamics import (
    HARDWARE_CATEGORY_OPTIONS,
    ReferenceProduct,
    append_products_to_cartridge_reference,
    build_hardware_sales_result,
    category_label_to_level,
    product_level_to_category,
)
from features.reference_update import (
    batch_add_clients_to_reference,
    batch_add_products_to_reference,
)
# Акцент для суммы ДЗ и числа контрагентов в заголовках (рядом с жирным текстом)
_HEADER_METRIC_VALUE_COLOR = "#1565c0"

_REFERENCE_ADDITIONS_LOG_KEY = "reference_additions_log"
# Ключи session_state для полей над блоками загрузки (неделя сбрасывается в app при новой загрузке).
CLIENT_BLOCK_WEEK_INPUT_KEY = "client_block_week_number_input"
EXCISE_LIQUID_PCS_INPUT_KEY = "excise_liquid_pcs_input"
SPEC_ORDERS_COUNT_INPUT_KEY = "spec_orders_count_input"
EXCISE_LIQUID_MARGIN_MULTIPLIER = 4.25

AI_REPORT_VERSION = "2026-07-16-v2"

AI_REPORT_SPEC_CATEGORY_LABELS = [
    "ОЭС 2 мл, шт.",
    "ОЭС 10 мл, шт.",
    "Жидкость 25 мл, шт.",
    "Pod-системы, шт.",
    "Расходники, шт.",
    "Кальянная продукция, шт.",
    "в т.ч. Уголь, шт.",
    "в т.ч. БКС/ТКС, шт.",
    "Никотиновые паучи, шт.",
    "Прочие товары, шт.",
]

AI_REPORT_TRADITION_CATEGORY_ROWS: list[tuple[str, str]] = [
    (
        "Одноразовые электронные сигареты ( 2 мл ) Традиция",
        "ОЭС 2 мл, шт.",
    ),
    (
        "Одноразовые электронные сигареты ( 10 мл ) Традиция",
        "ОЭС 10 мл, шт.",
    ),
    ("Никотиновые паучи Традиция", "Никотиновые паучи, шт."),
    ("в т.ч. Уголь, шт.", "в т.ч. Уголь, шт."),
]


def get_excise_liquid_margin_deduction() -> float:
    """Сумма, вычитаемая из общей маржи Спец.розницы: шт. акцизной жидкости × 4,25."""
    pcs = float(st.session_state.get(EXCISE_LIQUID_PCS_INPUT_KEY, 0) or 0)
    return pcs * EXCISE_LIQUID_MARGIN_MULTIPLIER


def get_spec_orders_count() -> int:
    """Кол-во заказов Спец. розницы из поля ввода над блоками РНП."""
    raw = st.session_state.get(SPEC_ORDERS_COUNT_INPUT_KEY, 0)
    try:
        return max(0, int(float(raw or 0)))
    except (TypeError, ValueError):
        return 0


def format_spec_orders_count() -> str:
    """Форматирование кол-ва заказов для таблиц отчётов."""
    return _format_quantity_compact(get_spec_orders_count())


def get_client_block_week_number(fallback: int) -> int:
    """Номер актуальной недели из поля ввода; если пусто — fallback из данных."""
    raw = st.session_state.get(CLIENT_BLOCK_WEEK_INPUT_KEY)
    if raw is None:
        return fallback
    if isinstance(raw, str):
        stripped = raw.strip()
        if not stripped:
            return fallback
        try:
            week = int(stripped)
        except ValueError:
            return fallback
    else:
        try:
            week = int(raw)
        except (TypeError, ValueError):
            return fallback
    if week < 1 or week > 53:
        return fallback
    return week


def render_global_rnp_inputs(default_excise_liquid: int = 0) -> None:
    """Поля «Актуальная неделя», «Акцизной жидкости шт.» и «Кол-во заказов» над блоками РНП."""
    st.markdown(
        f"""
        <style>
        .st-key-{CLIENT_BLOCK_WEEK_INPUT_KEY},
        .st-key-{EXCISE_LIQUID_PCS_INPUT_KEY},
        .st-key-{SPEC_ORDERS_COUNT_INPUT_KEY} {{
            max-width: 9.75rem;
            min-width: 9.75rem;
            width: 9.75rem;
        }}
        .st-key-{CLIENT_BLOCK_WEEK_INPUT_KEY} [data-testid="stTextInput"],
        .st-key-{CLIENT_BLOCK_WEEK_INPUT_KEY} [data-testid="stNumberInput"],
        .st-key-{EXCISE_LIQUID_PCS_INPUT_KEY} [data-testid="stNumberInput"],
        .st-key-{SPEC_ORDERS_COUNT_INPUT_KEY} [data-testid="stNumberInput"] {{
            max-width: 9.75rem;
            min-width: 9.75rem;
            width: 9.75rem;
        }}
        .st-key-{CLIENT_BLOCK_WEEK_INPUT_KEY} input,
        .st-key-{EXCISE_LIQUID_PCS_INPUT_KEY} input,
        .st-key-{SPEC_ORDERS_COUNT_INPUT_KEY} input {{
            max-width: 9.75rem;
            min-width: 9.75rem;
            width: 9.75rem;
        }}
        .rnp-global-inputs-label {{
            font-weight: 700;
            font-size: 0.95rem;
            line-height: 1.25;
            margin: 0 0 0.2rem 0;
            white-space: nowrap;
        }}
        .rnp-excise-deduction-hint {{
            font-size: 0.78rem;
            color: #1565c0;
            font-style: italic;
            font-weight: 400;
            margin: 0.45rem 0 0 0;
            white-space: nowrap;
            line-height: 1.35;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )

    _week_col, _excise_col, _orders_col, _spacer_col = st.columns(
        [1.75, 1.75, 1.75, 4.75], gap="large"
    )

    with _week_col:
        st.markdown(
            '<p class="rnp-global-inputs-label">Актуальная неделя:</p>',
            unsafe_allow_html=True,
        )
        if CLIENT_BLOCK_WEEK_INPUT_KEY not in st.session_state:
            st.session_state[CLIENT_BLOCK_WEEK_INPUT_KEY] = ""
        st.text_input(
            "Актуальная неделя",
            key=CLIENT_BLOCK_WEEK_INPUT_KEY,
            label_visibility="collapsed",
            placeholder="",
            help=(
                "Номер отчётной недели (1–53). Влияет на клиентский блок и имя Excel-отчёта. "
                "Если поле пустое, используется неделя из загруженных продаж."
            ),
        )

    with _excise_col:
        st.markdown(
            '<p class="rnp-global-inputs-label">Акцизной жидкости шт.:</p>',
            unsafe_allow_html=True,
        )
        st.number_input(
            "Акцизной жидкости шт.",
            min_value=0,
            value=default_excise_liquid,
            step=1,
            key=EXCISE_LIQUID_PCS_INPUT_KEY,
            label_visibility="collapsed",
            help=(
                "Умножается на 4,25 — полученная сумма вычитается только из общей маржи "
                "Спец.розницы (группы подразделений не меняются)."
            ),
        )

    with _orders_col:
        st.markdown(
            '<p class="rnp-global-inputs-label">Кол-во заказов:</p>',
            unsafe_allow_html=True,
        )
        st.number_input(
            "Кол-во заказов",
            min_value=0,
            value=0,
            step=1,
            key=SPEC_ORDERS_COUNT_INPUT_KEY,
            label_visibility="collapsed",
            help=(
                "Количество заказов B2B Спец. розницы. Подставляется в ИИ-отчёт "
                "и в блок «Общий РНП» (строка «Кол-во заказов»)."
            ),
        )

    with _spacer_col:
        st.empty()

    _deduction_display = _format_money(get_excise_liquid_margin_deduction())
    st.markdown(
        (
            '<p class="rnp-excise-deduction-hint">'
            f"Вычтено из МД {_deduction_display}</p>"
        ),
        unsafe_allow_html=True,
    )


def render_special_retail_dashboard(
    sales_df: pd.DataFrame,
    contractors_df: pd.DataFrame,
    categories_df: pd.DataFrame,
    category_order_df: pd.DataFrame | None = None,
    orders_df: pd.DataFrame | None = None,
    turnover_90_df: pd.DataFrame | None = None,
    turnover_7_df: pd.DataFrame | None = None,
    receivables_df: pd.DataFrame | None = None,
    cash_inflow_df: pd.DataFrame | None = None,
    hardware_levels_df: pd.DataFrame | None = None,
) -> None:
    merged_df, new_clients, unmatched_products = prepare_dataset(
        sales_df=sales_df,
        contractors_df=contractors_df,
        categories_df=categories_df,
        category_order_df=category_order_df,
    )
    spec_df = merged_df[
        merged_df["Подразделение"].isin(SPECIAL_RETAIL_SUBDIVISIONS)
    ]
    tradition_df = merged_df[merged_df["Подразделение"] == "Традиция"]
    target_client_name = 'ООО "Айса"'
    target_client_sales_df = _prepare_target_client_sales_df(
        sales_df=sales_df,
        categories_df=categories_df,
        target_client=target_client_name,
        category_order_df=category_order_df,
    )

    if "show_rnp_block" not in st.session_state:
        st.session_state.show_rnp_block = False
    if "show_general_rnp_block" not in st.session_state:
        st.session_state.show_general_rnp_block = False
    if "show_ai_rnp_block" not in st.session_state:
        st.session_state.show_ai_rnp_block = False

    if spec_df.empty:
        st.info("Нет данных по Спец. рознице.")
        return

    _excel_default_week = _client_block_default_week_from_spec(spec_df)
    _excel_week_num = get_client_block_week_number(_excel_default_week)
    excel_bytes = _build_full_report_excel(
        sales_df=sales_df,
        orders_df=orders_df,
        turnover_90_df=turnover_90_df,
        turnover_7_df=turnover_7_df,
        receivables_df=receivables_df,
        cash_inflow_df=cash_inflow_df,
        hardware_levels_df=hardware_levels_df,
        contractors_fallback_df=contractors_df,
        categories_fallback_df=categories_df,
        category_order_fallback_df=category_order_df,
    )
    _excel_file_name = f"РНП Спец. розница и Традиция — неделя {_excel_week_num}.xlsx"
    st.markdown(
        """
        <style>
        div[data-testid="stButton"] button[kind="tertiary"] {
            background-color: #0b2e6b !important;
            color: #ffffff !important;
            border: 1px solid #0b2e6b !important;
            font-weight: 800 !important;
            font-size: 1.05rem !important;
            border-radius: 10px !important;
            min-height: 44px !important;
            padding: 0.5rem 1rem !important;
            justify-content: flex-start !important;
            text-align: left !important;
        }
        div[data-testid="stButton"] button[kind="tertiary"]:hover,
        div[data-testid="stButton"] button[kind="tertiary"]:active,
        div[data-testid="stButton"] button[kind="tertiary"]:focus,
        div[data-testid="stButton"] button[kind="tertiary"]:focus-visible {
            background-color: #082554 !important;
            border-color: #082554 !important;
            color: #ffffff !important;
            box-shadow: none !important;
            outline: none !important;
        }
        div[data-testid="stButton"] button[kind="secondary"] {
            background-color: #1f5d35 !important;
            color: #ffffff !important;
            border: 1px solid #1f5d35 !important;
            font-weight: 800 !important;
            font-size: 1.05rem !important;
            border-radius: 10px !important;
            min-height: 44px !important;
            padding: 0.5rem 1rem !important;
            justify-content: flex-start !important;
            text-align: left !important;
        }
        div[data-testid="stButton"] button[kind="secondary"]:hover {
            background-color: #17472a !important;
            border-color: #17472a !important;
        }
        .st-key-toggle_ai_rnp_block_btn button {
            background-color: #b56a1a !important;
            color: #ffffff !important;
            border: 1px solid #b56a1a !important;
            font-weight: 800 !important;
            font-size: 1.05rem !important;
            border-radius: 10px !important;
            min-height: 44px !important;
            padding: 0.5rem 1rem !important;
            justify-content: flex-start !important;
            text-align: left !important;
        }
        .st-key-toggle_ai_rnp_block_btn button:hover,
        .st-key-toggle_ai_rnp_block_btn button:active,
        .st-key-toggle_ai_rnp_block_btn button:focus,
        .st-key-toggle_ai_rnp_block_btn button:focus-visible {
            background-color: #955716 !important;
            border-color: #955716 !important;
            color: #ffffff !important;
            box-shadow: none !important;
            outline: none !important;
        }
        div[data-testid="stDownloadButton"] button {
            background: transparent !important;
            color: inherit !important;
            border: 1px solid rgba(128, 128, 128, 0.4) !important;
            font-weight: 700 !important;
        }
        div[data-testid="stDownloadButton"] button:hover {
            background: rgba(128, 128, 128, 0.08) !important;
            border-color: rgba(128, 128, 128, 0.65) !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    if new_clients and unmatched_products:
        st.info("🆕 Обнаружены новые клиенты и товары, распределите их в справочнике ниже.")
    elif new_clients:
        st.info("🆕 Обнаружены новые клиенты, распределите их в справочнике ниже.")
    elif unmatched_products:
        st.info("🆕 Обнаружены новые товары, распределите их в справочнике ниже.")

    _render_quick_reference_update(
        new_clients=new_clients,
        unmatched_products=unmatched_products,
        contractors_df=contractors_df,
        categories_df=categories_df,
        receivables_df=receivables_df,
    )
    st.markdown("---")

    render_global_rnp_inputs()
    st.markdown("")

    toggle_label = (
        "▼ РНП B2B Спец.розница и Традиция (нажмите, чтобы свернуть)"
        if st.session_state.show_rnp_block
        else "▶ РНП B2B Спец.розница и Традиция (нажмите, чтобы развернуть)"
    )
    col_toggle, col_download = st.columns([1.35, 1], gap="small")
    with col_toggle:
        if st.button(
            toggle_label,
            key="toggle_rnp_block_btn",
            type="tertiary",
            use_container_width=True,
        ):
            st.session_state.show_rnp_block = not st.session_state.show_rnp_block
            st.rerun()
    with col_download:
        if excel_bytes is not None:
            st.download_button(
                "Скачать РНП отчёт в Excel",
                data=excel_bytes,
                file_name=_excel_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="secondary",
                use_container_width=True,
            )

    if st.session_state.show_rnp_block:
        st.markdown(
            "<h2 style='color:#8e44ad; margin-bottom:0.2rem;'>Спец. розница</h2>",
            unsafe_allow_html=True,
        )

        render_section(
            title="",
            dataframe=spec_df,
            categories_df=categories_df,
            category_order_df=category_order_df,
            use_expander=False,
            aggregates={"Минская область": MINSK_REGION_COMPONENTS},
            split_finance_categories=True,
            show_turnover=False,
            overall_margin_adjustment=get_excise_liquid_margin_deduction(),
        )

        allowed_subdivisions = set(spec_df["Подразделение"].dropna().astype(str))
        st.markdown("---")
        col_turnover, col_dz_spec, col_factor = st.columns([1, 1, 1], gap="small")
        with col_turnover:
            render_turnover_block(
                dataframe=spec_df,
                categories_df=categories_df,
                turnover_90_df=turnover_90_df,
                turnover_7_df=turnover_7_df,
                visible_rows=3,
                category_order_df=category_order_df,
            )
        with col_dz_spec:
            dz_table, dz_total, dz_error = _build_dz_spec_table(
                receivables_df=receivables_df
            )
            if dz_error:
                st.markdown("**ДЗ Спец розница**")
                st.info(dz_error)
            else:
                # Только phrasing-контент (без <p>), иначе вложенность ломает вертикальное выравнивание с соседними колонками
                st.markdown(
                    (
                        "<strong>ДЗ Спец розница - "
                        f'<span style="color:{_HEADER_METRIC_VALUE_COLOR};font-weight:700;">'
                        f"{_format_money(dz_total)}</span></strong>"
                    ),
                    unsafe_allow_html=True,
                )
                st.dataframe(
                    dz_table,
                    use_container_width=True,
                    hide_index=True,
                    height=_table_height_from_rows(3),
                    column_config={
                        "Контрагент": st.column_config.TextColumn("Контрагент"),
                        "ДЗ": st.column_config.TextColumn("ДЗ"),
                    },
                )
        with col_factor:
            render_factor_analysis(
                sales_df=spec_df,
                contractors_df=contractors_df,
                category_order_df=category_order_df,
            )

        st.markdown("---")
        _init_reference_additions_log()
        _client_count = _count_clients(spec_df)
        _client_count_display = _format_quantity(float(_client_count)) or "0"
        _default_week_num = _client_block_default_week_from_spec(spec_df)
        st.markdown(
            (
                "<div style='margin-top:6px;'><strong>Клиентский блок - кол-во контрагентов "
                f'<span style="color:{_HEADER_METRIC_VALUE_COLOR};font-weight:700;font-size:1.22em;">'
                f"{_client_count_display}</span></strong></div>"
            ),
            unsafe_allow_html=True,
        )
        _week_for_table = get_client_block_week_number(_default_week_num)
        spec_for_client_block = _spec_df_for_client_block_table(spec_df, _week_for_table)
        client_block_df = _build_client_block_table(spec_for_client_block)
        _client_height = _table_height_from_rows(4)
        st.dataframe(
            client_block_df,
            use_container_width=True,
            hide_index=True,
            height=_client_height,
            column_config={
                col: st.column_config.TextColumn(col)
                for col in client_block_df.columns
            },
        )

        st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)
        _panel_height = _table_height_from_rows(4)
        col_hardware, col_ref_log = st.columns([1.05, 1], gap="medium")
        with col_hardware:
            _render_hardware_sales_dynamics_panel(
                hardware_levels_df,
                table_height=_panel_height,
            )
        with col_ref_log:
            st.markdown(
                "<div style='margin-top:6px;'><strong>Добавлено в справочник</strong></div>",
                unsafe_allow_html=True,
            )
            _ref_log_df = _reference_additions_log_dataframe()
            if _ref_log_df.empty:
                st.caption(
                    "Пока нет записей. После нажатия «Обновить справочники и пересчитать отчёт» "
                    "в блоке быстрого добавления строки появятся здесь (только в этой сессии браузера)."
                )
            else:
                st.dataframe(
                    _ref_log_df,
                    use_container_width=True,
                    hide_index=True,
                    height=_panel_height,
                    column_config={
                        col: st.column_config.TextColumn(col)
                        for col in _ref_log_df.columns
                    },
                )

        st.markdown("---")
        with st.container():
            st.markdown(
                "<h2 style='color:#8e44ad; margin-bottom:0.2rem;'>Традиция</h2>",
                unsafe_allow_html=True,
            )
            st.markdown("<div style='height:14px;'></div>", unsafe_allow_html=True)
            if tradition_df.empty:
                st.info("Нет данных по Традиции.")
            else:
                tradition_table = _build_tradition_table_with_dz(
                    tradition_df=tradition_df,
                    receivables_df=receivables_df,
                    category_order_df=category_order_df,
                )
                st.dataframe(
                    tradition_table,
                    use_container_width=True,
                    hide_index=True,
                    height=290,
                    column_config={
                        col: st.column_config.TextColumn(col)
                        for col in tradition_table.columns
                        if col != "Показатель"
                    },
                )

    st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)
    general_toggle_label = (
        "▼ Общий РНП B2B Спец. розница и Традиция (нажмите, чтобы свернуть)"
        if st.session_state.show_general_rnp_block
        else "▶ Общий РНП B2B Спец. розница и Традиция (нажмите, чтобы развернуть)"
    )
    col_general_toggle, col_general_spacer = st.columns([1.35, 1], gap="small")
    with col_general_toggle:
        if st.button(
            general_toggle_label,
            key="toggle_general_rnp_block_btn",
            type="secondary",
            use_container_width=True,
        ):
            st.session_state.show_general_rnp_block = (
                not st.session_state.show_general_rnp_block
            )
            st.rerun()
    with col_general_spacer:
        st.empty()

    if st.session_state.show_general_rnp_block:
        general_table = _build_general_rnp_summary_table(
            spec_df=spec_df,
            tradition_df=tradition_df,
            target_client_df=target_client_sales_df,
            receivables_df=receivables_df,
            cash_inflow_df=cash_inflow_df,
            category_order_df=category_order_df,
        )
        st.dataframe(
            general_table,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Общие": st.column_config.TextColumn("Общие"),
                'ООО "Айса"': st.column_config.TextColumn('ООО "Айса"'),
            },
        )

    st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)
    ai_toggle_label = (
        "▼ ИИ отчёт B2B Спец.розница и Традиция (нажмите, чтобы свернуть)"
        if st.session_state.show_ai_rnp_block
        else "▶ ИИ отчёт B2B Спец.розница и Традиция (нажмите, чтобы развернуть)"
    )
    col_ai_toggle, col_ai_spacer = st.columns([1.35, 1], gap="small")
    with col_ai_toggle:
        if st.button(
            ai_toggle_label,
            key="toggle_ai_rnp_block_btn",
            type="primary",
            use_container_width=True,
        ):
            st.session_state.show_ai_rnp_block = not st.session_state.show_ai_rnp_block
            st.rerun()
    with col_ai_spacer:
        st.empty()

    if st.session_state.show_ai_rnp_block:
        st.caption(f"Версия шаблона ИИ-отчёта: {AI_REPORT_VERSION}")
        ai_table = build_ai_report_table(
            spec_df=spec_df,
            tradition_df=tradition_df,
            receivables_df=receivables_df,
            category_order_df=category_order_df,
        )
        st.dataframe(
            ai_table,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Значение": st.column_config.TextColumn("Значение"),
            },
        )


def _render_hardware_sales_dynamics_panel(
    hardware_levels_df: pd.DataFrame | None,
    table_height: int,
) -> None:
    """Таблица «Динамика продаж под-систем и расходников»."""
    st.markdown(
        (
            "<div style='margin-top:6px;'>"
            "<strong>Динамика продаж под-систем и расходников</strong>"
            "</div>"
        ),
        unsafe_allow_html=True,
    )
    try:
        cartridge_ref_df = load_reference(REF_SALES_POD_CARTRIDGE)
    except Exception as exc:  # noqa: BLE001
        st.warning(
            f"Справочник не найден: {get_reference_label(REF_SALES_POD_CARTRIDGE)}. "
            "Добавьте лист Sales_pod_cartridge в Google Sheets или файл "
            "Sales_pod_cartridge.xlsx в data/reference."
        )
        st.error(
            f"Не удалось загрузить справочник "
            f"{get_reference_label(REF_SALES_POD_CARTRIDGE)}: {exc}"
        )
        return

    try:
        hardware_result = build_hardware_sales_result(
            reference_df=cartridge_ref_df,
            levels_df=hardware_levels_df,
        )
    except ValueError as exc:
        st.error(str(exc))
        hardware_result = build_hardware_sales_result(
            reference_df=cartridge_ref_df,
            levels_df=None,
        )

    candidates = hardware_result.candidates_for_reference
    if candidates and hardware_levels_df is not None:
        with st.expander(
            f"Новые товары для справочника ({len(candidates)})",
            expanded=True,
        ):
            header_name_col, header_cat_col = st.columns([3.2, 1])
            with header_name_col:
                st.markdown("**Товар**")
            with header_cat_col:
                st.markdown("**Категория**")

            products_to_add: list[ReferenceProduct] = []
            for idx, item in enumerate(candidates):
                name_col, cat_col = st.columns([3.2, 1])
                default_category = product_level_to_category(item.level)
                with name_col:
                    st.markdown(item.name)
                with cat_col:
                    selected_category = st.selectbox(
                        "Категория",
                        HARDWARE_CATEGORY_OPTIONS,
                        index=HARDWARE_CATEGORY_OPTIONS.index(default_category),
                        key=f"hardware_candidate_category_{idx}",
                        label_visibility="collapsed",
                    )
                products_to_add.append(
                    ReferenceProduct(
                        name=item.name,
                        level=category_label_to_level(selected_category),
                    )
                )

            if st.button(
                f"Добавить {len(candidates)} товар(ов) в справочник Google Sheets",
                key="hardware_add_to_reference_btn",
                type="primary",
            ):
                try:
                    fresh_ref_df = load_reference(REF_SALES_POD_CARTRIDGE)
                    start_len = len(fresh_ref_df)
                    updated_ref, added_names = append_products_to_cartridge_reference(
                        fresh_ref_df,
                        products_to_add,
                    )
                    if not added_names:
                        st.info("Все найденные товары уже есть в справочнике.")
                    else:
                        new_rows = updated_ref.iloc[start_len:].to_dict("records")
                        append_reference_rows(
                            REF_SALES_POD_CARTRIDGE,
                            new_rows,
                        )
                        for product in products_to_add:
                            if product.name not in added_names:
                                continue
                            _append_reference_addition(
                                entry_type="Товар железа B2B",
                                element=product.name,
                                reference=get_reference_title(REF_SALES_POD_CARTRIDGE),
                                distribution=product_level_to_category(product.level),
                            )
                        st.success(
                            "Добавлено в справочник: " + ", ".join(added_names)
                        )
                        st.rerun()
                except Exception as exc:  # noqa: BLE001
                    st.error(
                        f"Не удалось обновить справочник "
                        f"{get_reference_label(REF_SALES_POD_CARTRIDGE)}: {exc}"
                    )

    hardware_table = hardware_result.table
    if hardware_table.empty:
        st.info("В справочнике нет товаров для отображения.")
        return

    hardware_table = hardware_table.copy()
    hardware_table["Продажи, шт."] = hardware_table["Продажи, шт."].map(
        lambda value: _format_quantity(float(value))
    )
    st.dataframe(
        hardware_table,
        use_container_width=True,
        hide_index=True,
        height=table_height,
        column_config={
            "Товар": st.column_config.TextColumn("Товар"),
            "Продажи, шт.": st.column_config.TextColumn("Продажи, шт."),
        },
    )


def _init_reference_additions_log() -> None:
    if _REFERENCE_ADDITIONS_LOG_KEY not in st.session_state:
        st.session_state[_REFERENCE_ADDITIONS_LOG_KEY] = []


def _append_reference_addition(
    entry_type: str,
    element: str,
    reference: str,
    distribution: str,
) -> None:
    """Журнал успешных записей в справочники (только текущая сессия Streamlit)."""
    _init_reference_additions_log()
    st.session_state[_REFERENCE_ADDITIONS_LOG_KEY].append(
        {
            "Время": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
            "Тип": entry_type,
            "Элемент": element,
            "Справочник": reference,
            "Распределение": distribution,
        }
    )


def _reference_additions_log_dataframe() -> pd.DataFrame:
    _init_reference_additions_log()
    cols = ["Время", "Тип", "Элемент", "Справочник", "Распределение"]
    rows: list = list(st.session_state.get(_REFERENCE_ADDITIONS_LOG_KEY, []))
    if not rows:
        return pd.DataFrame(columns=cols)
    df = pd.DataFrame(rows)
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    return df[cols].iloc[::-1].reset_index(drop=True)


def _render_quick_reference_update(
    new_clients: list[str],
    unmatched_products: list[tuple[str, str, str]],
    contractors_df: pd.DataFrame,
    categories_df: pd.DataFrame,
    receivables_df: pd.DataFrame | None,
) -> None:
    dz_pending = _collect_unassigned_dz_counterparties(receivables_df)
    if not new_clients and not unmatched_products and dz_pending.empty:
        return

    st.markdown("### Быстрое добавление в справочники")
    subdivision_options = sorted(
        contractors_df.get("Подразделение", pd.Series(dtype="string"))
        .dropna()
        .astype(str)
        .str.strip()
        .replace("", pd.NA)
        .dropna()
        .unique()
        .tolist()
    )
    category_options = sorted(
        categories_df.get(
            get_category_source_column(categories_df) or "Категория",
            pd.Series(dtype="string"),
        )
        .dropna()
        .astype(str)
        .str.strip()
        .replace("", pd.NA)
        .dropna()
        .unique()
        .tolist()
    )
    razrez_col = get_razrez_source_column(categories_df) or "Разрез"
    razrez_options = [""] + sorted(
        categories_df.get(razrez_col, pd.Series(dtype="string"))
        .fillna("")
        .astype(str)
        .str.strip()
        .loc[lambda s: s.ne("")]
        .unique()
        .tolist()
    )

    with st.form("quick_reference_update_form"):
        st.markdown(
            """
            <style>
            div[data-testid="stFormSubmitButton"] button {
                background-color: #2e7d32 !important;
                color: #ffffff !important;
                border: 1px solid #2e7d32 !important;
                font-weight: 700 !important;
            }
            div[data-testid="stFormSubmitButton"] button:hover {
                background-color: #256b2a !important;
                border-color: #256b2a !important;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )
        if new_clients:
            st.markdown(
                "<div style='font-size:20px; font-weight:700; text-transform:uppercase; margin:6px 0 10px 0;'>Новые клиенты:</div>",
                unsafe_allow_html=True,
            )
            header_client_col, header_subdivision_col = st.columns([1, 1.8], gap="small")
            with header_client_col:
                st.markdown("**Клиент**")
            with header_subdivision_col:
                sub_header_col, _, _ = st.columns([1.2, 1, 1], gap="small")
                sub_header_col.markdown("**Подразделение**")
            for client in new_clients:
                client_col, subdivision_col = st.columns([1, 1.8], gap="small")
                with client_col:
                    st.markdown(f"`{client}`")
                with subdivision_col:
                    sub_col, _, _ = st.columns([1.2, 1, 1], gap="small")
                    with sub_col:
                        st.selectbox(
                            f"Подразделение для клиента: {client}",
                            options=subdivision_options,
                            index=0 if subdivision_options else None,
                            key=f"new_client_subdivision_{client}",
                            placeholder="Выберите подразделение",
                            label_visibility="collapsed",
                        )

        if unmatched_products:
            st.markdown(
                "<div style='font-size:20px; font-weight:700; text-transform:uppercase; margin:10px 0 10px 0;'>Новые товары:</div>",
                unsafe_allow_html=True,
            )
            header_product_col, header_controls_col = st.columns([1, 1.8], gap="small")
            with header_product_col:
                st.markdown("**Товар**")
            with header_controls_col:
                c1, c2 = st.columns([1.2, 1], gap="small")
                c1.markdown("**Категория**")
                c2.markdown("**Разрез (необязательно)**")
            for idx, (prod1, prod2, prod3) in enumerate(unmatched_products):
                product_col, controls_col = st.columns([1, 1.8], gap="small")
                title = " / ".join(
                    [p for p in [prod1, prod2, prod3] if p and p != "__NONE__"]
                )
                with product_col:
                    st.markdown(f"`{title}`")

                with controls_col:
                    category_col, razrez_col_ui = st.columns([1.2, 1], gap="small")
                    with category_col:
                        st.selectbox(
                            f"Категория для товара #{idx + 1}",
                            options=category_options,
                            index=0 if category_options else None,
                            key=f"new_product_category_{idx}",
                            placeholder="Выберите категорию",
                            label_visibility="collapsed",
                        )
                    with razrez_col_ui:
                        st.selectbox(
                            f"Разрез для товара #{idx + 1}",
                            options=razrez_options,
                            key=f"new_product_razrez_{idx}",
                            label_visibility="collapsed",
                        )

        if not dz_pending.empty:
            st.markdown(
                "<div style='font-size:20px; font-weight:700; text-transform:uppercase; margin:10px 0 10px 0;'>Новые контрагенты с ДЗ:</div>",
                unsafe_allow_html=True,
            )
            h_name_col, h_dz_col, h_target_col = st.columns([1.4, 0.7, 0.9], gap="small")
            h_name_col.markdown("**Контрагент**")
            h_dz_col.markdown("**ДЗ**")
            h_target_col.markdown("**Куда добавить**")
            choices = ["спец розница", "традиция", "убрать"]
            for idx, row in dz_pending.reset_index(drop=True).iterrows():
                c1, c2, c3 = st.columns([1.4, 0.7, 0.9], gap="small")
                c1.markdown(f"`{row['Контрагент']}`")
                c2.markdown(_format_money(float(row["ДЗ"])))
                c3.selectbox(
                    f"Распределение нового контрагента ДЗ #{idx + 1}",
                    options=choices,
                    index=0,
                    key=f"dz_distribution_target_{idx}",
                    label_visibility="collapsed",
                )

        apply_clicked = st.form_submit_button(
            "Обновить справочники и пересчитать отчёт",
            type="primary",
        )

    if not apply_clicked:
        return

    clients_added = 0
    clients_updated = 0
    products_added = 0
    dz_added = 0
    failed_messages: list[str] = []
    references_changed = False

    if new_clients:
        client_items = [
            (
                client,
                st.session_state.get(f"new_client_subdivision_{client}", ""),
            )
            for client in new_clients
        ]
        for (client, selected_subdivision), (ok, message) in zip(
            client_items,
            batch_add_clients_to_reference(client_items),
            strict=True,
        ):
            if ok:
                references_changed = True
                if "Обновлён клиент" in message:
                    clients_updated += 1
                    _append_reference_addition(
                        "Клиент (обновлено)",
                        client,
                        "Контрагенты",
                        f"Подразделение: {selected_subdivision}",
                    )
                else:
                    clients_added += 1
                    _append_reference_addition(
                        "Клиент (новый)",
                        client,
                        "Контрагенты",
                        f"Подразделение: {selected_subdivision}",
                    )
            else:
                failed_messages.append(message)

    if unmatched_products:
        product_items = [
            (
                product_levels,
                st.session_state.get(f"new_product_category_{idx}", ""),
                st.session_state.get(f"new_product_razrez_{idx}", ""),
            )
            for idx, product_levels in enumerate(unmatched_products)
        ]
        for item, (ok, message) in zip(
            product_items,
            batch_add_products_to_reference(product_items),
            strict=True,
        ):
            product_levels, selected_category, selected_razrez = item
            if ok:
                references_changed = True
                products_added += 1
                p1, p2, p3 = product_levels
                product_title = " / ".join(
                    [p for p in (p1, p2, p3) if p and p != "__NONE__"]
                )
                dist_parts = [f"Категория: {selected_category}"]
                if str(selected_razrez).strip():
                    dist_parts.append(f"Разрез: {selected_razrez}")
                _append_reference_addition(
                    "Товар",
                    product_title,
                    "Категории товаров",
                    "; ".join(dist_parts),
                )
            else:
                failed_messages.append(message)

    if not dz_pending.empty:
        target_map = {
            "спец розница": REF_DZ_SPEC,
            "традиция": REF_DZ_TRAD,
            "убрать": REF_DZ_REMOVE,
        }
        dz_by_ref: dict[str, list[tuple[str, str]]] = {}
        for idx, row in dz_pending.reset_index(drop=True).iterrows():
            target = st.session_state.get(f"dz_distribution_target_{idx}", "спец розница")
            ref_key = target_map[target]
            dz_by_ref.setdefault(ref_key, []).append(
                (str(row["Контрагент"]), target)
            )
        for ref_key, entries in dz_by_ref.items():
            names = [name for name, _ in entries]
            for (counterparty, target), (ok, message) in zip(
                entries,
                _batch_append_counterparties_to_dz_reference(ref_key, names),
                strict=True,
            ):
                if ok and message != "Контрагент уже есть в справочнике.":
                    references_changed = True
                    dz_added += 1
                    _append_reference_addition(
                        "ДЗ контрагент",
                        counterparty,
                        get_reference_title(ref_key),
                        f"Справочник: {get_reference_label(ref_key)}; выбор: {target}",
                    )
                elif not ok:
                    failed_messages.append(message)

    if clients_added:
        st.success(f"Добавлено клиентов: {clients_added}")
    if clients_updated:
        st.success(f"Обновлено клиентов: {clients_updated}")
    if products_added:
        st.success(f"Добавлено товаров: {products_added}")
    if dz_added:
        st.success(f"Распределено контрагентов ДЗ: {dz_added}")
    for msg in failed_messages:
        st.warning(msg)

    if references_changed:
        _refresh_reference_session_state(
            contractors_updated=bool(clients_added or clients_updated),
            categories_updated=bool(products_added),
        )
        st.rerun()


def _render_dz_counterparties_distribution(
    receivables_df: pd.DataFrame | None,
) -> None:
    pending = _collect_unassigned_dz_counterparties(receivables_df)
    if pending.empty:
        return

    st.markdown("### Распределение новых контрагентов ДЗ")
    st.caption(
        "Показаны только контрагенты с ДЗ > 0, которых нет в справочниках: "
        "ДЗ_Спец_розница, ДЗ_Традиция, ДЗ_Убрать."
    )

    choices = ["спец розница", "традиция", "убрать"]
    with st.form("dz_counterparties_distribution_form"):
        header_name_col, header_dz_col, header_target_col = st.columns(
            [1.4, 0.7, 0.9], gap="small"
        )
        header_name_col.markdown("**Контрагент**")
        header_dz_col.markdown("**ДЗ**")
        header_target_col.markdown("**Куда добавить**")

        for idx, row in pending.reset_index(drop=True).iterrows():
            c1, c2, c3 = st.columns([1.4, 0.7, 0.9], gap="small")
            c1.markdown(f"`{row['Контрагент']}`")
            c2.markdown(_format_money(float(row["ДЗ"])))
            c3.selectbox(
                f"Распределение контрагента ДЗ #{idx + 1}",
                options=choices,
                index=0,
                key=f"dz_standalone_target_{idx}",
                label_visibility="collapsed",
            )

        apply_clicked = st.form_submit_button(
            "Обновить справочники ДЗ",
            type="primary",
        )

    if not apply_clicked:
        return

    target_map = {
        "спец розница": REF_DZ_SPEC,
        "традиция": REF_DZ_TRAD,
        "убрать": REF_DZ_REMOVE,
    }
    added = 0
    failed_messages: list[str] = []
    pending_rows = pending.reset_index(drop=True)
    dz_by_ref: dict[str, list[tuple[str, str]]] = {}
    for idx, row in pending_rows.iterrows():
        target = st.session_state.get(f"dz_standalone_target_{idx}", "спец розница")
        ref_key = target_map[target]
        dz_by_ref.setdefault(ref_key, []).append((str(row["Контрагент"]), target))

    for ref_key, entries in dz_by_ref.items():
        names = [name for name, _ in entries]
        for (counterparty, target), (ok, message) in zip(
            entries,
            _batch_append_counterparties_to_dz_reference(ref_key, names),
            strict=True,
        ):
            if ok and message != "Контрагент уже есть в справочнике.":
                added += 1
                _append_reference_addition(
                    "ДЗ контрагент",
                    counterparty,
                    get_reference_title(ref_key),
                    f"Справочник: {get_reference_label(ref_key)}; выбор: {target}",
                )
            elif not ok:
                failed_messages.append(message)

    if added:
        st.success(f"Распределено контрагентов ДЗ: {added}")
    for msg in failed_messages:
        st.warning(msg)
    if added:
        st.rerun()


def _collect_unassigned_dz_counterparties(
    receivables_df: pd.DataFrame | None,
) -> pd.DataFrame:
    if receivables_df is None or receivables_df.empty or receivables_df.shape[1] < 7:
        return pd.DataFrame(columns=["Контрагент", "ДЗ", "key"])

    source = receivables_df.iloc[:, [1, 6]].copy()
    source.columns = ["Контрагент", "ДЗ"]
    source["Контрагент"] = source["Контрагент"].fillna("").astype(str).str.strip()
    source["ДЗ"] = pd.to_numeric(source["ДЗ"], errors="coerce")
    source = source[source["Контрагент"].ne("") & source["ДЗ"].notna() & source["ДЗ"].gt(0)]
    if source.empty:
        return pd.DataFrame(columns=["Контрагент", "ДЗ", "key"])

    source["key"] = source["Контрагент"].map(_normalize_client_name)
    grouped = source.groupby("key", as_index=False)["ДЗ"].sum()
    first_names = (
        source.drop_duplicates("key")[["key", "Контрагент"]]
        .reset_index(drop=True)
    )
    grouped = grouped.merge(first_names, on="key", how="left")

    existing_keys = (
        _load_dz_reference_keys(REF_DZ_SPEC)
        | _load_dz_reference_keys(REF_DZ_TRAD)
        | _load_dz_reference_keys(REF_DZ_REMOVE)
    )
    result = grouped[~grouped["key"].isin(existing_keys)].copy()
    return result[["Контрагент", "ДЗ", "key"]]


def _load_dz_reference_keys(ref_key: str) -> set[str]:
    try:
        df = load_reference(ref_key)
    except Exception:  # noqa: BLE001
        return set()
    if "Контрагент" not in df.columns:
        return set()
    return set(
        df["Контрагент"]
        .fillna("")
        .astype(str)
        .str.strip()
        .replace("", pd.NA)
        .dropna()
        .map(_normalize_client_name)
        .tolist()
    )


def _refresh_reference_session_state(
    *,
    contractors_updated: bool = False,
    categories_updated: bool = False,
) -> None:
    """Синхронизирует session_state с кэшем справочников после записи."""
    if categories_updated:
        try:
            st.session_state.categories_df = load_reference(REF_CATEGORIES)
        except Exception:  # noqa: BLE001
            pass
    if contractors_updated:
        try:
            st.session_state.contractors_df = load_reference(REF_CONTRACTORS)
        except Exception:  # noqa: BLE001
            pass


def _batch_append_counterparties_to_dz_reference(
    ref_key: str,
    counterparties: list[str],
) -> list[tuple[bool, str]]:
    label = get_reference_label(ref_key)
    if not counterparties:
        return []

    try:
        df = load_reference(ref_key)
    except Exception as exc:  # noqa: BLE001
        message = f"Не удалось прочитать {label}: {exc}"
        return [(False, message) for _ in counterparties]

    if "Контрагент" not in df.columns:
        message = f"В справочнике {label} нет столбца «Контрагент»."
        return [(False, message) for _ in counterparties]

    norm_existing = set(
        df["Контрагент"]
        .fillna("")
        .astype(str)
        .str.strip()
        .map(_normalize_client_name)
        .tolist()
    )
    results: list[tuple[bool, str]] = []
    rows_to_append: list[dict[str, object]] = []

    for counterparty in counterparties:
        value = str(counterparty).strip()
        if not value:
            results.append((False, "Пустое имя контрагента."))
            continue
        norm_value = _normalize_client_name(value)
        if norm_value in norm_existing:
            results.append((True, "Контрагент уже есть в справочнике."))
            continue

        new_row = {col: "" for col in df.columns}
        new_row["Контрагент"] = value
        rows_to_append.append(new_row)
        norm_existing.add(norm_value)
        results.append(
            (True, f"Добавлен контрагент в {get_reference_title(ref_key)}")
        )

    if rows_to_append:
        try:
            append_reference_rows(ref_key, rows_to_append)
        except Exception as exc:  # noqa: BLE001
            message = f"Не удалось записать {label}: {exc}"
            return [
                (False, message) if ok else (ok, msg)
                for ok, msg in results
            ]

    return results


def _append_counterparty_to_dz_reference(ref_key: str, counterparty: str) -> tuple[bool, str]:
    results = _batch_append_counterparties_to_dz_reference(ref_key, [counterparty])
    return results[0] if results else (False, "Нет данных для добавления.")


def _build_dz_spec_table(
    receivables_df: pd.DataFrame | None,
) -> tuple[pd.DataFrame, float, str | None]:
    try:
        dz_ref_df = load_reference(REF_DZ_SPEC)
    except Exception as exc:  # noqa: BLE001
        return pd.DataFrame(), 0.0, f"Не удалось прочитать справочник ДЗ: {exc}"
    if "Контрагент" not in dz_ref_df.columns:
        return pd.DataFrame(), 0.0, "В справочнике ДЗ_Спец_розница нет столбца «Контрагент»."
    if receivables_df is None or receivables_df.empty:
        return pd.DataFrame(), 0.0, "Загрузите файл «Дебиторская задолженность (62 счёт)»."
    if receivables_df.shape[1] < 7:
        return pd.DataFrame(), 0.0, "В файле ДЗ недостаточно столбцов (ожидался 7-й столбец с ДЗ)."

    source = receivables_df.iloc[:, [1, 6]].copy()
    source.columns = ["Контрагент", "ДЗ"]
    source["Контрагент"] = source["Контрагент"].fillna("").astype(str).str.strip()
    source["ДЗ"] = pd.to_numeric(source["ДЗ"], errors="coerce")
    source = source[
        source["Контрагент"].ne("") & source["ДЗ"].notna()
    ]
    if source.empty:
        ref_list = (
            dz_ref_df["Контрагент"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", pd.NA)
            .dropna()
            .tolist()
        )
        empty_table = pd.DataFrame(
            {"Контрагент": ref_list, "ДЗ": [""] * len(ref_list)}
        )
        return empty_table, 0.0, None

    source["key"] = source["Контрагент"].map(_normalize_client_name)
    grouped = source.groupby("key", as_index=False)["ДЗ"].sum()

    ref = dz_ref_df[["Контрагент"]].copy()
    ref["Контрагент"] = ref["Контрагент"].fillna("").astype(str).str.strip()
    ref = ref[ref["Контрагент"].ne("")]
    ref["key"] = ref["Контрагент"].map(_normalize_client_name)

    merged = ref.merge(grouped, on="key", how="left")
    merged["ДЗ"] = merged["ДЗ"].fillna(0.0)
    total_value = float(merged["ДЗ"].sum())
    table = pd.DataFrame(
        {
            "Контрагент": merged["Контрагент"],
            "ДЗ": merged["ДЗ"].map(_format_money),
        }
    )
    return table, total_value, None


def _build_tradition_table_with_dz(
    tradition_df: pd.DataFrame,
    receivables_df: pd.DataFrame | None,
    category_order_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    subdivisions = ["Традиция"]
    tradition_order = load_category_order_list(category_order_df, COL_TRADITION_RNP)
    metrics_table = build_financial_metrics_table(
        tradition_df,
        subdivisions=subdivisions,
        include_overall=False,
        aggregates={},
    )
    category_table = build_category_table(
        tradition_df,
        category_order=tradition_order,
        subdivisions=subdivisions,
        include_overall=False,
        aggregates={},
    )

    metrics_part = metrics_table.rename(columns={"Метрика": "Показатель"}).copy()
    categories_part = category_table.rename(columns={"Категория": "Показатель"}).copy()
    value_columns = [col for col in metrics_part.columns if col != "Показатель"]
    if not value_columns:
        value_columns = [col for col in categories_part.columns if col != "Показатель"]
    target_columns = ["Показатель"] + value_columns
    metrics_part = metrics_part.reindex(columns=target_columns, fill_value="")
    categories_part = categories_part.reindex(columns=target_columns, fill_value="")

    dz_total, _ = _calc_dz_total_by_reference(
        reference_key=REF_DZ_TRAD,
        receivables_df=receivables_df,
    )
    dz_row = {"Показатель": "ДЗ на конец недели"}
    for col in value_columns:
        dz_row[col] = ""
    if value_columns:
        dz_row[value_columns[0]] = _format_money(dz_total)
    spacer_row = {"Показатель": ""}
    for col in value_columns:
        spacer_row[col] = ""

    return pd.concat(
        [metrics_part, pd.DataFrame([dz_row]), pd.DataFrame([spacer_row]), categories_part],
        ignore_index=True,
    )


def _calc_dz_total_by_reference(
    reference_key: str,
    receivables_df: pd.DataFrame | None,
) -> tuple[float, str | None]:
    if receivables_df is None or receivables_df.empty:
        return 0.0, "no_receivables"
    try:
        ref_df = load_reference(reference_key)
    except Exception:  # noqa: BLE001
        return 0.0, "read_error"
    if "Контрагент" not in ref_df.columns:
        return 0.0, "bad_reference"
    if receivables_df.shape[1] < 7:
        return 0.0, "bad_receivables"

    source = receivables_df.iloc[:, [1, 6]].copy()
    source.columns = ["Контрагент", "ДЗ"]
    source["Контрагент"] = source["Контрагент"].fillna("").astype(str).str.strip()
    source["ДЗ"] = pd.to_numeric(source["ДЗ"], errors="coerce")
    source = source[source["Контрагент"].ne("") & source["ДЗ"].notna()]
    if source.empty:
        return 0.0, None
    source["key"] = source["Контрагент"].map(_normalize_client_name)
    grouped = source.groupby("key", as_index=False)["ДЗ"].sum()

    ref = ref_df[["Контрагент"]].copy()
    ref["Контрагент"] = ref["Контрагент"].fillna("").astype(str).str.strip()
    ref = ref[ref["Контрагент"].ne("")]
    ref["key"] = ref["Контрагент"].map(_normalize_client_name)

    merged = ref.merge(grouped, on="key", how="left")
    merged["ДЗ"] = merged["ДЗ"].fillna(0.0)
    return float(merged["ДЗ"].sum()), None


def _client_block_default_week_from_spec(spec_df: pd.DataFrame) -> int:
    """Номер недели по умолчанию для виджета: максимум из данных или 1."""
    week_col = _find_week_column_name(spec_df)
    if not week_col or week_col not in spec_df.columns:
        return 1
    series = pd.to_numeric(spec_df[week_col], errors="coerce").dropna()
    if series.empty:
        return 1
    return int(round(float(series.max())))


def _find_week_column_name(df: pd.DataFrame) -> str | None:
    for col in df.columns:
        if str(col).strip().lower() == "неделя":
            return str(col)
    return None


def _spec_df_for_client_block_table(spec_df: pd.DataFrame, week_value: int) -> pd.DataFrame:
    """Копия Спец. розницы: во всех строках столбец «Неделя» = week_value."""
    out = spec_df.copy()
    try:
        week_val = int(week_value)
    except (TypeError, ValueError):
        week_val = 1
    week_col = _find_week_column_name(out)
    if week_col:
        out[week_col] = week_val
    else:
        out["Неделя"] = week_val
    return out


def _format_week_cell(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, str) and not value.strip():
        return ""
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value).strip()
    if pd.isna(num):
        return ""
    if num == int(num):
        return str(int(num))
    return str(value).strip()


def _display_product_level(series: pd.Series) -> pd.Series:
    out = series.fillna("").astype(str).str.strip()
    out = out.mask(out.eq("__NONE__"), "")
    out = out.mask(out.str.lower().eq("nan"), "")
    return out


def _build_client_block_table(spec_df: pd.DataFrame) -> pd.DataFrame:
    """Детализация продаж Спец. розницы для отображения (как в файле «Продажи»)."""
    out_columns = [
        "Клиент",
        "Группа контрагентов",
        "Товар ур.2",
        "Товар ур.3",
        "Неделя",
        "Продажи с НДС",
        "Маржа",
        "Количество",
    ]
    if spec_df.empty:
        return pd.DataFrame(columns=out_columns)

    df = spec_df.copy()
    sort_cols: list[str] = []
    if "Клиент" in df.columns:
        sort_cols.append("Клиент")
    if "Группа контрагентов" in df.columns:
        sort_cols.append("Группа контрагентов")
    for col in ("Товар ур.2", "Товар ур.3"):
        if col in df.columns:
            sort_cols.append(col)
    if sort_cols:
        df = df.sort_values(by=sort_cols, ascending=True, kind="stable")

    clients = (
        df["Клиент"].fillna("").astype(str).str.strip()
        if "Клиент" in df.columns
        else pd.Series([""] * len(df), index=df.index)
    )
    groups_raw = (
        df["Группа контрагентов"].fillna("").astype(str).str.strip()
        if "Группа контрагентов" in df.columns
        else pd.Series([""] * len(df), index=df.index)
    )
    groups = groups_raw.replace("", "-")

    p2 = _display_product_level(df["Товар ур.2"]) if "Товар ур.2" in df.columns else pd.Series([""] * len(df), index=df.index)
    p3 = _display_product_level(df["Товар ур.3"]) if "Товар ур.3" in df.columns else pd.Series([""] * len(df), index=df.index)

    week_col = _find_week_column_name(df)
    if week_col:
        week_series = df[week_col].map(_format_week_cell)
    else:
        week_series = pd.Series([""] * len(df), index=df.index, dtype=object)

    sales_num = pd.to_numeric(df.get("Продажи с НДС", 0.0), errors="coerce").fillna(0.0)
    margin_num = pd.to_numeric(df.get("Маржа", 0.0), errors="coerce").fillna(0.0)
    qty_num = pd.to_numeric(df.get("Количество", 0.0), errors="coerce").fillna(0.0)

    return pd.DataFrame(
        {
            "Клиент": clients,
            "Группа контрагентов": groups,
            "Товар ур.2": p2,
            "Товар ур.3": p3,
            "Неделя": week_series,
            "Продажи с НДС": sales_num.map(lambda x: _format_money(float(x))),
            "Маржа": margin_num.map(lambda x: _format_money(float(x))),
            "Количество": qty_num.map(lambda x: _format_quantity(float(x))),
        }
    )


def _build_client_block_export_table(spec_df: pd.DataFrame) -> pd.DataFrame:
    """Таблица для листа Excel «Клиентский блок».

    Первая колонка — «Подразделение». В «Клиент» к имени добавлено подразделение (через « — »).
    «Категория товара» — из «Категория агрег.». Колонки «Неделя» нет. Числа без текстового форматирования.
    """
    out_columns = [
        "Подразделение",
        "Клиент",
        "Группа контрагентов",
        "Категория товара",
        "Товар ур.2",
        "Товар ур.3",
        "Продажи с НДС",
        "Маржа",
        "Количество",
    ]
    if spec_df.empty:
        return pd.DataFrame(columns=out_columns)

    df = spec_df.copy()
    sort_cols: list[str] = []
    if "Подразделение" in df.columns:
        sort_cols.append("Подразделение")
    if "Клиент" in df.columns:
        sort_cols.append("Клиент")
    if "Группа контрагентов" in df.columns:
        sort_cols.append("Группа контрагентов")
    if "Категория агрег." in df.columns:
        sort_cols.append("Категория агрег.")
    for col in ("Товар ур.2", "Товар ур.3"):
        if col in df.columns:
            sort_cols.append(col)
    if sort_cols:
        df = df.sort_values(by=sort_cols, ascending=True, kind="stable")

    clients = (
        df["Клиент"].fillna("").astype(str).str.strip()
        if "Клиент" in df.columns
        else pd.Series([""] * len(df), index=df.index)
    )
    subdiv = (
        df["Подразделение"].fillna("").astype(str).str.strip()
        if "Подразделение" in df.columns
        else pd.Series([""] * len(df), index=df.index)
    )
    groups_raw = (
        df["Группа контрагентов"].fillna("").astype(str).str.strip()
        if "Группа контрагентов" in df.columns
        else pd.Series([""] * len(df), index=df.index)
    )
    groups = groups_raw.replace("", "-")
    category = (
        df["Категория агрег."].fillna("").astype(str).str.strip()
        if "Категория агрег." in df.columns
        else pd.Series([""] * len(df), index=df.index)
    )

    p2 = _display_product_level(df["Товар ур.2"]) if "Товар ур.2" in df.columns else pd.Series([""] * len(df), index=df.index)
    p3 = _display_product_level(df["Товар ур.3"]) if "Товар ур.3" in df.columns else pd.Series([""] * len(df), index=df.index)

    sales_num = pd.to_numeric(df.get("Продажи с НДС", 0.0), errors="coerce")
    margin_num = pd.to_numeric(df.get("Маржа", 0.0), errors="coerce")
    qty_num = pd.to_numeric(df.get("Количество", 0.0), errors="coerce")

    sub_nonempty = subdiv.astype(str).str.len() > 0
    clients_annotated = clients.astype(str)
    clients_annotated = clients_annotated.where(
        ~sub_nonempty, clients_annotated + " — " + subdiv.astype(str)
    )

    return pd.DataFrame(
        {
            "Подразделение": subdiv,
            "Клиент": clients_annotated,
            "Группа контрагентов": groups,
            "Категория товара": category,
            "Товар ур.2": p2,
            "Товар ур.3": p3,
            "Продажи с НДС": sales_num,
            "Маржа": margin_num,
            "Количество": qty_num,
        }
    )


def _vertical_metrics_lookup(table: pd.DataFrame) -> dict[str, str]:
    if table.empty:
        return {}
    result: dict[str, str] = {}
    for _, row in table.iterrows():
        label = str(row.get("Показатель", "")).strip()
        if not label:
            continue
        result[label] = str(row.get("Значение", "")).strip()
    return result


def _lookup_category_row_value(categories: dict[str, str], label: str) -> str:
    if label in categories:
        return categories[label]
    target_key = _label_key(label)
    for key, value in categories.items():
        if _label_key(key) == target_key:
            return str(value)
    return ""


def build_ai_report_table(
    spec_df: pd.DataFrame,
    tradition_df: pd.DataFrame,
    receivables_df: pd.DataFrame | None,
    category_order_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    spec_order = load_category_order_list(category_order_df, COL_SPEC_RNP)
    tradition_order = load_category_order_list(category_order_df, COL_TRADITION_RNP)
    margin_adjustment = get_excise_liquid_margin_deduction()

    spec_finance = _vertical_metrics_lookup(
        build_financial_metrics_vertical_table(
            spec_df,
            subdivisions=[],
            include_overall=True,
            aggregates={},
            format_money=_format_money_compact,
            overall_margin_adjustment=margin_adjustment,
        )
    )
    tradition_finance = _vertical_metrics_lookup(
        build_financial_metrics_vertical_table(
            tradition_df,
            subdivisions=[],
            include_overall=True,
            aggregates={},
            format_money=_format_money_compact,
        )
    )

    spec_categories = extract_category_row_values(
        spec_df,
        spec_order,
        format_value=_format_quantity_compact,
    )
    tradition_categories = extract_category_row_values(
        tradition_df,
        tradition_order,
        format_value=_format_quantity_compact,
    )

    rtrade_mask = (
        spec_df["Клиент"].fillna("").astype(str).str.strip().isin(RTRADE_CLIENTS)
        if "Клиент" in spec_df.columns
        else pd.Series(False, index=spec_df.index)
    )
    rtrade_sales = _safe_sum(spec_df.loc[rtrade_mask], "Продажи с НДС")
    rtrade_margin = _safe_sum(spec_df.loc[rtrade_mask], "Маржа")
    dz_spec_total, _ = _calc_dz_total_by_reference(REF_DZ_SPEC, receivables_df)
    dz_trad_total, _ = _calc_dz_total_by_reference(REF_DZ_TRAD, receivables_df)

    rows: list[dict[str, str]] = [
        {"Показатель": "Заказы", "Значение": format_spec_orders_count()},
        {
            "Показатель": "Кол-во клиентов сделавших заказ B2B Спец.розница",
            "Значение": str(_count_clients(spec_df)),
        },
        {
            "Показатель": "Продажи с НДС B2B Спец.розница",
            "Значение": spec_finance.get("Продажи с НДС", ""),
        },
        {
            "Показатель": "Маржа B2B Спец.розница",
            "Значение": spec_finance.get("Маржа", ""),
        },
        {
            "Показатель": "% Маржи B2B Спец.розница",
            "Значение": spec_finance.get("% МД", ""),
        },
        {
            "Показатель": "Продажи с НДС Ртрейд",
            "Значение": _format_money_compact(rtrade_sales),
        },
        {
            "Показатель": "Маржа Ртрейд",
            "Значение": _format_money_compact(rtrade_margin),
        },
        {
            "Показатель": "Дебиторская задолженность B2B Спец.розница",
            "Значение": _format_money_compact(dz_spec_total),
        },
    ]

    for category_label in AI_REPORT_SPEC_CATEGORY_LABELS:
        rows.append(
            {
                "Показатель": category_label,
                "Значение": _lookup_category_row_value(spec_categories, category_label),
            }
        )

    rows.extend(
        [
            {
                "Показатель": "Продажи с НДС Традиция",
                "Значение": tradition_finance.get("Продажи с НДС", ""),
            },
            {
                "Показатель": "Маржа Традиция",
                "Значение": tradition_finance.get("Маржа", ""),
            },
            {
                "Показатель": "% Маржи Традиция",
                "Значение": tradition_finance.get("% МД", ""),
            },
            {
                "Показатель": "Дебиторская задолженность Традиция",
                "Значение": _format_money_compact(dz_trad_total),
            },
        ]
    )

    for display_label, lookup_label in AI_REPORT_TRADITION_CATEGORY_ROWS:
        rows.append(
            {
                "Показатель": display_label,
                "Значение": _lookup_category_row_value(
                    tradition_categories, lookup_label
                ),
            }
        )

    return pd.DataFrame(rows)


def _safe_sum(df: pd.DataFrame, col: str) -> float:
    if col not in df.columns or df.empty:
        return 0.0
    return float(pd.to_numeric(df[col], errors="coerce").fillna(0.0).sum())


def _build_general_rnp_summary_table(
    spec_df: pd.DataFrame,
    tradition_df: pd.DataFrame,
    target_client_df: pd.DataFrame,
    receivables_df: pd.DataFrame | None,
    cash_inflow_df: pd.DataFrame | None,
    category_order_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    general_spec_order = load_category_order_list(category_order_df, COL_GENERAL_SPEC)
    general_tradition_order = load_category_order_list(
        category_order_df, COL_GENERAL_TRADITION
    )
    spec_category_overall = extract_category_row_values(
        df=spec_df,
        order=general_spec_order,
        format_value=_format_quantity_compact,
    )
    spec_category_client = extract_category_row_values(
        df=target_client_df,
        order=general_spec_order,
        format_value=_format_quantity_compact,
    )
    spec_metrics_overall = _extract_metrics_rows(
        spec_df,
        format_money=_format_money_compact,
        overall_margin_adjustment=get_excise_liquid_margin_deduction(),
    )
    spec_metrics_client = _extract_metrics_rows(
        target_client_df, format_money=_format_money_compact
    )

    tradition_category_overall = extract_category_row_values(
        df=tradition_df,
        order=general_tradition_order,
        format_value=_format_quantity_compact,
    )
    tradition_metrics_overall = _extract_metrics_rows(
        tradition_df, format_money=_format_money_compact
    )

    dz_spec_total, _ = _calc_dz_total_by_reference(REF_DZ_SPEC, receivables_df)
    dz_trad_total, _ = _calc_dz_total_by_reference(REF_DZ_TRAD, receivables_df)
    cash_inflow_trad_total = _calc_cash_inflow_total_for_tradition(cash_inflow_df)

    rows: list[dict[str, str]] = [
        {
            "Блок": "Спец розница",
            "Показатель": "Кол-во заказов",
            "Общие": format_spec_orders_count(),
            'ООО "Айса"': "",
        },
        {
            "Блок": "",
            "Показатель": "Кол-во клиентов",
            "Общие": str(_count_clients(spec_df)),
            'ООО "Айса"': "",
        },
    ]
    for category in spec_category_overall.keys():
        rows.append(
            {
                "Блок": "",
                "Показатель": category,
                "Общие": spec_category_overall.get(category, ""),
                'ООО "Айса"': spec_category_client.get(category, ""),
            }
        )
    for metric_name in ("Выручка", "МД", "МД%"):
        rows.append(
            {
                "Блок": "",
                "Показатель": metric_name,
                "Общие": spec_metrics_overall.get(metric_name, ""),
                'ООО "Айса"': spec_metrics_client.get(metric_name, ""),
            }
        )
    rows.extend(
        [
            {
                "Блок": "",
                "Показатель": "ДЗ спец розница",
                "Общие": _format_money_compact(dz_spec_total),
                'ООО "Айса"': "",
            },
            {
                "Блок": "",
                "Показатель": "просроченная ДЗ",
                "Общие": "",
                'ООО "Айса"': "",
            },
            {"Блок": "", "Показатель": "", "Общие": "", 'ООО "Айса"': ""},
        ]
    )
    first_trad = True
    for category in tradition_category_overall.keys():
        rows.append(
            {
                "Блок": "Традиция" if first_trad else "",
                "Показатель": category,
                "Общие": tradition_category_overall.get(category, ""),
                'ООО "Айса"': "",
            }
        )
        first_trad = False
    rows.append(
        {
            "Блок": "",
            "Показатель": "ДЗ Традиция",
            "Общие": _format_money_compact(dz_trad_total),
            'ООО "Айса"': "",
        }
    )
    for metric_name in ("Выручка", "МД", "МД%"):
        rows.append(
            {
                "Блок": "",
                "Показатель": metric_name,
                "Общие": tradition_metrics_overall.get(metric_name, ""),
                'ООО "Айса"': "",
            }
        )
    rows.append(
        {
            "Блок": "",
            "Показатель": "Поступление ДС",
            "Общие": _format_money_compact(cash_inflow_trad_total),
            'ООО "Айса"': "",
        }
    )
    return pd.DataFrame(rows)


def _calc_cash_inflow_total_for_tradition(cash_inflow_df: pd.DataFrame | None) -> float:
    if cash_inflow_df is None or cash_inflow_df.empty:
        return 0.0
    if cash_inflow_df.shape[1] < 11:
        return 0.0

    tradition_keys = _load_dz_reference_keys(REF_DZ_TRAD)
    if not tradition_keys:
        return 0.0

    source = cash_inflow_df.iloc[:, [10, 8]].copy()
    source.columns = ["Контрагент", "Сумма"]
    source["Контрагент"] = source["Контрагент"].fillna("").astype(str).str.strip()
    source["Сумма"] = pd.to_numeric(source["Сумма"], errors="coerce").fillna(0.0)
    source = source[source["Контрагент"].ne("")]
    if source.empty:
        return 0.0

    source["key"] = source["Контрагент"].map(_normalize_client_name)
    matched = source[source["key"].isin(tradition_keys)]
    if matched.empty:
        return 0.0
    return float(matched["Сумма"].sum())


def _extract_metrics_rows(
    df: pd.DataFrame,
    *,
    format_money: Callable[[float | int | None], str] | None = None,
    overall_margin_adjustment: float = 0.0,
) -> dict[str, str]:
    table = build_financial_metrics_vertical_table(
        df=df,
        subdivisions=[],
        include_overall=True,
        aggregates={},
        format_money=format_money,
        overall_margin_adjustment=overall_margin_adjustment,
    )
    if table.empty:
        return {"Выручка": "", "МД": "", "МД%": ""}
    mapped_names = {
        "Продажи с НДС": "Выручка",
        "Маржа": "МД",
        "% МД": "МД%",
    }
    result: dict[str, str] = {"Выручка": "", "МД": "", "МД%": ""}
    for _, row in table.iterrows():
        source_name = str(row.get("Показатель", "")).strip()
        target_name = mapped_names.get(source_name)
        if target_name:
            result[target_name] = str(row.get("Значение", ""))
    return result


def _filter_by_client(df: pd.DataFrame, target_client: str) -> pd.DataFrame:
    if df.empty or "Клиент" not in df.columns:
        return df.iloc[0:0].copy()
    normalized_target = _normalize_client_name(target_client)
    target_tokens = set(_tokenize_client_name(normalized_target))
    clients = df["Клиент"].fillna("").astype(str).map(_normalize_client_name)

    mask_exact = clients.eq(normalized_target)
    mask_tokens = clients.map(
        lambda value: _is_target_client_name(value, target_tokens=target_tokens)
    )
    return df.loc[mask_exact | mask_tokens].copy()


def _prepare_target_client_sales_df(
    sales_df: pd.DataFrame,
    categories_df: pd.DataFrame,
    target_client: str,
    category_order_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    if sales_df is None or sales_df.empty:
        return pd.DataFrame()

    df = sales_df.copy()
    for numeric_col in ("Количество", "Продажи с НДС", "Маржа"):
        if numeric_col not in df.columns:
            df[numeric_col] = 0.0
        df[numeric_col] = pd.to_numeric(df[numeric_col], errors="coerce").fillna(0.0)

    if "Категория агрег." not in df.columns or "Разрез" not in df.columns:
        df = _merge_categories(df, categories_df, category_order_df)
    else:
        df["Категория агрег."] = df["Категория агрег."].fillna("Прочие товары, шт.:")
        df["Разрез"] = df["Разрез"].fillna("")

    match_columns = []
    if "Клиент" in df.columns:
        match_columns.append("Клиент")
    if "Группа контрагентов" in df.columns:
        match_columns.append("Группа контрагентов")
    if not match_columns:
        return df.iloc[0:0].copy()

    masks = [_filter_by_client(df[[col]].rename(columns={col: "Клиент"}), target_client).index for col in match_columns]
    if not masks:
        return df.iloc[0:0].copy()
    matched_index = masks[0]
    for idx in masks[1:]:
        matched_index = matched_index.union(idx)
    return df.loc[matched_index].copy()


def _normalize_client_name(value: str) -> str:
    normalized = (
        str(value)
        .strip()
        .replace("«", '"')
        .replace("»", '"')
        .replace("“", '"')
        .replace("”", '"')
        .replace("„", '"')
    )
    return " ".join(normalized.split()).lower()


def _tokenize_client_name(value: str) -> list[str]:
    cleaned = "".join(
        ch if (ch.isalnum() or ch.isspace()) else " "
        for ch in value.lower()
    )
    return [part for part in cleaned.split() if part]


def _is_target_client_name(value: str, target_tokens: set[str]) -> bool:
    if not value.strip():
        return False
    value_tokens = set(_tokenize_client_name(value))
    if not value_tokens:
        return False
    if target_tokens.issubset(value_tokens):
        return True
    # Фолбэк: если в названии явно есть ключевая часть клиента.
    return "айса" in value_tokens


def _extract_orders_week_number(
    orders_df: pd.DataFrame | None,
    contractors_df: pd.DataFrame,
    allowed_subdivisions: set[str],
) -> str:
    if orders_df is None or orders_df.empty:
        return "Без недели"
    prepared_df, _ = _prepare_orders_dataset(
        orders_df=orders_df,
        contractors_df=contractors_df,
        allowed_subdivisions=allowed_subdivisions,
    )
    if prepared_df.empty or "Неделя" not in prepared_df.columns:
        return "Без недели"
    week_series = pd.to_numeric(prepared_df["Неделя"], errors="coerce").dropna()
    if week_series.empty:
        return "Без недели"
    return str(int(round(float(week_series.max()))))


def _read_latest_reference(ref_key: str, fallback_df: pd.DataFrame) -> pd.DataFrame:
    try:
        return load_reference(ref_key)
    except Exception:  # noqa: BLE001
        return fallback_df


def _build_hardware_dynamics_export_table(
    hardware_levels_df: pd.DataFrame | None,
) -> pd.DataFrame:
    """Таблица «Динамика продаж под-систем и расходников» для Excel."""
    try:
        cartridge_ref_df = load_reference(REF_SALES_POD_CARTRIDGE)
    except Exception:  # noqa: BLE001
        cartridge_ref_df = pd.DataFrame()

    try:
        hardware_result = build_hardware_sales_result(
            reference_df=cartridge_ref_df,
            levels_df=hardware_levels_df,
        )
    except ValueError:
        hardware_result = build_hardware_sales_result(
            reference_df=cartridge_ref_df,
            levels_df=None,
        )

    table = hardware_result.table.copy()
    if table.empty:
        return table

    table["Продажи, шт."] = pd.to_numeric(
        table["Продажи, шт."], errors="coerce"
    ).fillna(0.0)
    return table


def _build_full_report_excel(
    sales_df: pd.DataFrame,
    orders_df: pd.DataFrame | None,
    turnover_90_df: pd.DataFrame | None,
    turnover_7_df: pd.DataFrame | None,
    receivables_df: pd.DataFrame | None,
    cash_inflow_df: pd.DataFrame | None,
    contractors_fallback_df: pd.DataFrame,
    categories_fallback_df: pd.DataFrame,
    category_order_fallback_df: pd.DataFrame | None = None,
    hardware_levels_df: pd.DataFrame | None = None,
) -> bytes | None:
    contractors_df = _read_latest_reference(REF_CONTRACTORS, contractors_fallback_df)
    categories_df = _read_latest_reference(REF_CATEGORIES, categories_fallback_df)
    category_order_df = _read_latest_reference(
        REF_CATEGORY_ORDER,
        category_order_fallback_df if category_order_fallback_df is not None else pd.DataFrame(),
    )

    merged_df, _, _ = prepare_dataset(
        sales_df=sales_df,
        contractors_df=contractors_df,
        categories_df=categories_df,
        category_order_df=category_order_df,
    )
    spec_df = merged_df[
        merged_df["Подразделение"].isin(SPECIAL_RETAIL_SUBDIVISIONS)
    ].copy()
    tradition_df = merged_df[merged_df["Подразделение"] == "Традиция"].copy()
    if spec_df.empty:
        return None

    aggregates = {"Минская область": MINSK_REGION_COMPONENTS}
    subdivisions = _collect_subdivisions_for_export(spec_df, aggregates)
    spec_category_order = load_category_order_list(category_order_df, COL_SPEC_RNP)

    financial_overall = build_financial_metrics_vertical_table(
        spec_df,
        [],
        include_overall=True,
        aggregates=aggregates,
        overall_margin_adjustment=get_excise_liquid_margin_deduction(),
    )
    _, dz_spec_total, _ = _build_dz_spec_table(receivables_df)
    financial_overall = _append_dz_metric_row(financial_overall, dz_spec_total)
    financial_subdiv = build_financial_metrics_vertical_table(
        spec_df, subdivisions, include_overall=False, aggregates=aggregates
    )
    category_overall = build_category_vertical_table(
        spec_df,
        category_order=spec_category_order,
        subdivisions=[],
        include_overall=True,
        aggregates=aggregates,
    )
    category_subdiv = build_category_vertical_table(
        spec_df,
        category_order=spec_category_order,
        subdivisions=subdivisions,
        include_overall=False,
        aggregates=aggregates,
        spacer_between_groups=True,
    )
    turnover_table = _build_turnover_export_table(
        spec_df=spec_df,
        categories_df=categories_df,
        turnover_90_df=turnover_90_df,
        turnover_7_df=turnover_7_df,
        category_order_df=category_order_df,
    )
    factor_table = _build_factor_export_table(
        spec_df, categories_df, category_order_df
    )
    orders_summary_table, orders_category_table = _build_orders_export_tables(
        orders_df=orders_df,
        contractors_df=contractors_df,
        categories_df=categories_df,
        allowed_subdivisions=set(spec_df["Подразделение"].dropna().astype(str)),
        category_order_df=category_order_df,
    )
    dz_spec_table, _, _ = _build_dz_spec_table(receivables_df)
    tradition_table = _build_tradition_export_table(
        tradition_df=tradition_df,
        receivables_df=receivables_df,
        category_order_df=category_order_df,
    )
    hardware_dynamics_table = _build_hardware_dynamics_export_table(
        hardware_levels_df
    )
    client_block_export = _build_client_block_export_table(spec_df)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        finance_sheet = _stack_tables_with_titles(
            [
                ("Финансы — Общие", financial_overall),
                ("Финансы — Подразделения", financial_subdiv),
            ]
        )
        categories_turnover_sheet = _stack_tables_with_titles(
            [
                ("Категории — Общие", category_overall),
                ("Категории — Подразделения", category_subdiv),
                ("Оборачиваемость", turnover_table),
            ]
        )
        orders_sheet = _stack_tables_with_titles(
            [
                ("Заказы — Сводка", orders_summary_table),
                ("Заказы — Категории", orders_category_table),
            ]
        )
        sheets_to_write = [
            ("Финансы", finance_sheet),
            ("Категории_Оборач.", categories_turnover_sheet),
            ("ДЗ Спец розница", dz_spec_table),
            ("Факторный анализ", factor_table),
            ("Традиция", tradition_table),
            ("Динамика железа", hardware_dynamics_table),
        ]
        if _has_export_data(orders_sheet):
            sheets_to_write.append(("Заказы", orders_sheet))
        written_count = 0
        for sheet_name, sheet_df in sheets_to_write:
            if sheet_name == "Динамика железа":
                if sheet_df is not None and not sheet_df.empty:
                    _write_sheet(writer, sheet_name, sheet_df)
                    written_count += 1
            elif _has_export_data(sheet_df):
                _write_sheet(writer, sheet_name, sheet_df)
                written_count += 1
        if _has_export_data(client_block_export):
            _write_client_block_sheet_with_autofilter(
                writer, "Клиентский блок", client_block_export
            )
            written_count += 1
        if written_count:
            _autosize_columns(writer)
        else:
            _write_sheet(writer, "Отчёт", pd.DataFrame([{"Статус": "Нет данных для выгрузки"}]))
            _autosize_columns(writer)

    return output.getvalue()


def _build_orders_export_tables(
    orders_df: pd.DataFrame | None,
    contractors_df: pd.DataFrame,
    categories_df: pd.DataFrame,
    allowed_subdivisions: set[str],
    category_order_df: pd.DataFrame | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if orders_df is None or orders_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    prepared_df, _ = _prepare_orders_dataset(
        orders_df=orders_df,
        contractors_df=contractors_df,
        allowed_subdivisions=allowed_subdivisions,
    )
    if prepared_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    last_week = prepared_df["Неделя"].max()
    if pd.isna(last_week):
        return pd.DataFrame(), pd.DataFrame()

    last_week_df = prepared_df[prepared_df["Неделя"] == last_week]
    last_week_num = int(round(float(last_week)))
    summary = pd.DataFrame(
        [
            {"Показатель": "Контрагентов с начала цикла", "Значение": _count_clients(prepared_df)},
            {"Показатель": f"Контрагентов на неделе {last_week_num}", "Значение": _count_clients(last_week_df)},
            {"Показатель": f"Наполненность (шт.), неделя {last_week_num}", "Значение": int(round(last_week_df["Количество"].sum()))},
        ]
    )
    category_table = calculate_orders_category_metrics(
        prepared_orders_df=prepared_df,
        categories_df=categories_df,
        last_week_value=last_week,
        category_order_df=category_order_df,
    )
    if category_table is None:
        category_table = pd.DataFrame()
    return summary, category_table


def _append_dz_metric_row(financial_table: pd.DataFrame, dz_total: float) -> pd.DataFrame:
    if financial_table is None or financial_table.empty:
        return financial_table
    if "Показатель" not in financial_table.columns:
        return financial_table

    md_idx = financial_table.index[
        financial_table["Показатель"].fillna("").astype(str).str.strip().eq("% МД")
    ].tolist()
    if not md_idx:
        return financial_table

    insert_at = md_idx[0] + 1
    dz_row = {col: "" for col in financial_table.columns}
    if "Группа" in dz_row:
        dz_row["Группа"] = ""
    dz_row["Показатель"] = "ДЗ на конец недели"
    if "Значение" in dz_row:
        dz_row["Значение"] = _format_money(dz_total)

    upper = financial_table.iloc[:insert_at]
    lower = financial_table.iloc[insert_at:]
    return pd.concat([upper, pd.DataFrame([dz_row]), lower], ignore_index=True)


def _build_tradition_export_table(
    tradition_df: pd.DataFrame,
    receivables_df: pd.DataFrame | None,
    category_order_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    if tradition_df.empty:
        return pd.DataFrame()
    tradition_order = load_category_order_list(category_order_df, COL_TRADITION_RNP)
    metrics_raw = build_financial_metrics_vertical_table(
        tradition_df, subdivisions=[], include_overall=True, aggregates={}
    )
    metrics_table = metrics_raw.rename(
        columns={"Группа": "Показатель", "Показатель": "Метрика", "Значение": "Традиция"}
    )[["Метрика", "Традиция"]].rename(columns={"Метрика": "Показатель"})
    dz_trad_total, _ = _calc_dz_total_by_reference(REF_DZ_TRAD, receivables_df)
    dz_row = pd.DataFrame(
        [{"Показатель": "ДЗ на конец недели", "Традиция": _format_money(dz_trad_total)}]
    )
    spacer_row = pd.DataFrame([{"Показатель": "", "Традиция": ""}])
    metrics_with_dz = pd.concat(
        [metrics_table, dz_row, spacer_row],
        ignore_index=True,
    )

    categories_raw = build_category_vertical_table(
        tradition_df,
        category_order=tradition_order,
        subdivisions=[],
        include_overall=True,
        aggregates={},
    )
    category_table = categories_raw.rename(
        columns={"Категория": "Показатель", "Значение": "Традиция"}
    )[["Показатель", "Традиция"]]

    return _stack_tables_with_titles(
        [
            ("Традиция — Финансовые метрики", metrics_with_dz),
            ("Традиция — Продажи по категориям (шт.)", category_table),
        ]
    )


def _build_factor_export_table(
    spec_df: pd.DataFrame,
    categories_df: pd.DataFrame,
    category_order_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    prepared_df = _prepare_factor_base(spec_df)
    if prepared_df is None:
        merged = _merge_categories(spec_df.copy(), categories_df, category_order_df)
        prepared_df = _prepare_factor_base(merged)
    if prepared_df is None:
        return pd.DataFrame()
    order = load_category_order_list(category_order_df, COL_SPEC_RNP)
    segment_table = pd.DataFrame(_build_segment_rows(prepared_df))
    quantity_table = pd.DataFrame(
        _build_category_rows(prepared_df, value_column="Количество", category_order=order)
    )
    revenue_table = pd.DataFrame(
        _build_category_rows(prepared_df, value_column="Продажи с НДС", category_order=order)
    )
    return _stack_tables_with_titles(
        [
            ("Выручка по сегментам (Продажи с НДС)", segment_table),
            ("Продажи по категориям (шт.)", quantity_table),
            ("Выручка по категориям (Продажи с НДС)", revenue_table),
        ]
    )


def _build_turnover_export_table(
    spec_df: pd.DataFrame,
    categories_df: pd.DataFrame,
    turnover_90_df: pd.DataFrame | None,
    turnover_7_df: pd.DataFrame | None,
    category_order_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    clients_filter = (
        spec_df["Клиент"].dropna().astype(str).unique().tolist()
        if "Клиент" in spec_df.columns
        else None
    )
    table_90 = calculate_turnover_by_category(
        turnover_df=turnover_90_df,
        categories_df=categories_df,
        clients_filter=clients_filter,
        period_days=90,
        category_order_df=category_order_df,
    )
    table_7 = calculate_turnover_by_category(
        turnover_df=turnover_7_df,
        categories_df=categories_df,
        clients_filter=clients_filter,
        period_days=7,
        category_order_df=category_order_df,
    )
    turnover_order = load_category_order_list(category_order_df, COL_TURNOVER)
    return build_turnover_table(table_90, table_7, category_order=turnover_order)


def _collect_subdivisions_for_export(
    df: pd.DataFrame, aggregates: dict[str, list[str]]
) -> list[str]:
    available: set[str] = set()
    if "Подразделение" in df.columns:
        available.update(
            df["Подразделение"].dropna().astype(str).str.strip().unique().tolist()
        )
    available.update(aggregates.keys())
    ordered = [sub for sub in SPECIAL_RETAIL_SUBDIVISIONS if sub in available]
    others = sorted(available - set(ordered))
    return ordered + others


def _write_sheet(
    writer: pd.ExcelWriter, sheet_name: str, dataframe: pd.DataFrame
) -> None:
    dataframe.to_excel(writer, sheet_name=sheet_name, index=False)


def _write_client_block_sheet_with_autofilter(
    writer: pd.ExcelWriter, sheet_name: str, dataframe: pd.DataFrame
) -> None:
    """Записывает лист «Клиентский блок» и включает автофильтр по строке заголовков."""
    dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    ncols = max(1, int(dataframe.shape[1]))
    last_col = get_column_letter(ncols)
    nrows = max(1, int(len(dataframe)) + 1)
    ws.auto_filter.ref = f"A1:{last_col}{nrows}"


def _autosize_columns(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    section_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    header_font = Font(bold=True, color="1F1F1F")
    section_font = Font(bold=True, color="1F1F1F")

    for sheet in writer.book.worksheets:
        # Header row formatting
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Body formatting + section row highlight
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            first_value = row[0].value
            first_text = "" if first_value is None else str(first_value).strip()
            is_section_row = (
                first_text.startswith("Секция:")
                or first_text.startswith("Секция —")
                or first_text.startswith("Секция ")
                or first_text.startswith("Финансы —")
                or first_text.startswith("Категории —")
                or first_text.startswith("Заказы —")
                or first_text.startswith("Выручка по ")
                or first_text.startswith("Продажи по ")
                or first_text.startswith("Традиция —")
            )
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
                if is_section_row:
                    cell.fill = section_fill
                    cell.font = section_font

        for column_cells in sheet.columns:
            max_len = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            sheet.column_dimensions[column_letter].width = min(max_len + 2, 80)


def _stack_tables_with_titles(
    parts: list[tuple[str, pd.DataFrame]]
) -> pd.DataFrame:
    stacked: list[pd.DataFrame] = []
    valid_parts = [(title, table) for title, table in parts if table is not None and not table.empty]
    for idx, (title, table) in enumerate(valid_parts):
        title_row = pd.DataFrame([{"Секция": title}])
        stacked.append(title_row)
        stacked.append(table.reset_index(drop=True))
        if idx < len(valid_parts) - 1:
            stacked.append(pd.DataFrame([{}]))
            stacked.append(pd.DataFrame([{}]))
    if not stacked:
        return pd.DataFrame()
    return pd.concat(stacked, ignore_index=True, sort=False)


def _has_export_data(df: pd.DataFrame) -> bool:
    if df is None or df.empty:
        return False
    text = (
        df.fillna("")
        .astype(str)
        .apply(lambda s: s.str.strip())
    )
    non_empty_cells = text.ne("").sum().sum()
    if non_empty_cells == 0:
        return False
    return True